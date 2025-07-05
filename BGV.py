import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ----------------------------
# 🗕 Public Holidays
# ----------------------------
public_holidays = pd.to_datetime([
    "2025-01-26", "2025-08-15", "2025-10-02", "2025-12-25"
])

# ----------------------------
# 🧠 Helper Functions
# ----------------------------
def is_working_day(date):
    if date.weekday() == 6:
        return False
    if date.weekday() == 5:
        week = (date.day - 1) // 7 + 1
        if week in [2, 4]:
            return False
    if date in public_holidays:
        return False
    return True

def add_working_days(start_date, n):
    date = start_date
    count = 0
    while count < n:
        date += timedelta(days=1)
        if is_working_day(date):
            count += 1
    return date

def calculate_due(row):
    if pd.notnull(row['BGV_Reinitiated']):
        return add_working_days(row['BGV_Reinitiated'], 8)
    elif pd.notnull(row['BGV_Received On']):
        return add_working_days(row['BGV_Received On'], 15)
    return pd.NaT

def calculate_remarks(row):
    dispatch = row['BGV_Final Dispatch']
    due = row['Final TAT Due Date for Report']
    if pd.isnull(dispatch) or pd.isnull(due):
        return "Pending", ""
    diff = (dispatch - due).days
    if diff <= 0:
        return "Within TAT", ""
    return "Exceeded", f"{diff} days Deduction"

def process_report(df):
    for col in df.columns:
        if "Date" in col or "On" in col or "Reinitiated" in col or "Dispatch" in col:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    df['Final TAT Due Date for Report'] = df.apply(calculate_due, axis=1)
    df[['Remarks', 'Due Days']] = df.apply(lambda row: pd.Series(calculate_remarks(row)), axis=1)

    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%d-%b-%Y')

    return df

def style_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl', date_format='DD-MMM-YYYY') as writer:
        df.to_excel(writer, sheet_name='BGV_Report', index=False)
        sheet = writer.sheets['BGV_Report']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        center = Alignment(horizontal='center', vertical='center')

        for col_num, col in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        remarks_col = df.columns.get_loc("Remarks") + 1
        due_col = df.columns.get_loc("Due Days") + 1

        for row in range(2, sheet.max_row + 1):
            remark = sheet.cell(row=row, column=remarks_col).value
            color = None
            if remark == "Within TAT":
                color = green_fill
            elif remark == "Exceeded":
                color = red_fill
            elif remark == "Pending":
                color = yellow_fill

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if row % 2 == 0 and not color:
                    cell.fill = alt_fill
                if col in [remarks_col, due_col] and color:
                    cell.fill = color
                cell.alignment = center

        for col_num, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            sheet.column_dimensions[get_column_letter(col_num)].width = max_len + 2

    output.seek(0)
    return output

# ----------------------------
# 🌐 Streamlit UI
# ----------------------------
st.set_page_config("BGV Report Generator", layout="wide", page_icon="📊")
st.markdown("""
    <style>
        .main { background-color: #eef4fa; }
        .block-container { padding-top: 2rem; }
        .stButton>button {
            font-size: 16px;
            border-radius: 8px;
            background-color: #2e7bcf;
            color: white;
            padding: 8px 16px;
        }
        .stButton>button:hover {
            background-color: #1b5eaa;
        }
        .title-container {
            text-align: center;
            padding: 1rem;
        }
        .logo {
            height: 80px;
        }
    </style>
""", unsafe_allow_html=True)

with st.container():
    col1, col2, col3 = st.columns([1, 5, 1])
    with col2:
        st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/a/ae/Logo_BGV.svg/1280px-Logo_BGV.svg.png", width=120)
        st.markdown("<h2 style='text-align: center; color: #2e7bcf;'>📋 BGV Final TAT Report Generator</h2>", unsafe_allow_html=True)

# 📅 Sidebar Instructions
with st.sidebar:
    st.markdown("## 📋 Instructions")
    st.markdown("""
    1. **Download** the BGV Excel template.
    2. **Fill in** all required fields using **DD-MMM-YYYY** date format.
    3. **Do not** alter column headers.
    4. Upload to generate TAT summary.

    ---
    - 🟢 Green: Within TAT
    - 🔴 Red: Exceeded
    - 🟡 Yellow: Pending
    """)
    st.markdown("### 📧 Contact")
    st.info("For issues, contact MIS support team.")

# 📅 Template Download
with st.expander("⬇️ Download Excel Template", expanded=True):
    template_columns = [
        "Sl.No", "CandidateCode", "Candidate Name",
        "BWR_Date of Submission", "BWR_TAT Due On", "BWR_Reinitiated", "BWR_Date of Report Received",
        "BGV_Received On", "BGV_TAT Due On", "BGV_Reinitiated", "BGV_Final Dispatch"
    ]
    template_df = pd.DataFrame(columns=template_columns)
    buffer = io.BytesIO()
    template_df.to_excel(buffer, index=False)
    st.download_button("📄 Download Template", buffer.getvalue(), file_name="BGV_Template.xlsx")

# 📄 Upload + Report Generation
st.subheader("📄 Upload Filled Template")
uploaded_file = st.file_uploader("Upload the filled Excel file", type="xlsx")

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        missing_cols = [col for col in template_columns if col not in df.columns]
        if missing_cols:
            st.error(f"❌ Missing columns: {', '.join(missing_cols)}")
        else:
            result_df = process_report(df)
            st.success("✅ Report generated successfully!")

            st.subheader("🔍 Preview Report")
            st.dataframe(result_df, use_container_width=True)

            excel_data = style_excel(result_df)
            st.download_button("📁 Download Final Report", excel_data, file_name="BGV_Final_TAT_Report.xlsx")
    except Exception as e:
        st.error(f"⚠️ Error processing file: {e}")
